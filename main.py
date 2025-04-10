import os
import psycopg2
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from datetime import datetime
import sys
import copy

def main():
    # 1. Obtener credenciales de la base de datos
    db_name = os.environ.get('DB_NAME', 'semillas')
    db_user = os.environ.get('DB_USER', 'openerp')
    db_password = os.environ.get('DB_PASSWORD', '')
    db_host = os.environ.get('DB_HOST', '2.136.142.253')
    db_port = os.environ.get('DB_PORT', '5432')
    
    db_params = {
        'dbname': db_name,
        'user': db_user,
        'password': db_password,
        'host': db_host,
        'port': db_port
    }
    
    # 2. Generar el nombre del fichero con fecha y hora (timestamp)
    now = datetime.now()
    timestamp = now.strftime("%Y%m%d_%H%M%S")
    # Si no se ha definido la variable de entorno EXCEL_FILE_PATH, se genera con el timestamp
    file_path = os.environ.get('EXCEL_FILE_PATH', f'Productos_{timestamp}.xlsx')
    
    # 3. Consulta SQL para extraer la información de productos
    query = """
    SELECT 
        (SELECT name FROM res_company WHERE id = p.company_func_id) AS "Compañía",
        'Granada' AS "Sede",
        (SELECT name FROM product_category WHERE id = p.seccion) AS "Sección",
        (SELECT name FROM product_category WHERE id = p.familia) AS "Familia",
        (SELECT name FROM product_category WHERE id = p.subfamilia) AS "Subfamilia",
        p.default_code AS "Referencia",
        p.name AS "Nombre",
 
        CASE 
            WHEN (
                (SELECT COALESCE(SUM(product_qty), 0)
                 FROM stock_move 
                 WHERE product_id = p.id 
                   AND date <= CURRENT_DATE
                   AND location_dest_id IN (60, 80)
                   AND historico != 'yes'
                   AND state = 'done')
                -
                (SELECT COALESCE(SUM(product_qty), 0)
                 FROM stock_move 
                 WHERE product_id = p.id 
                   AND date <= CURRENT_DATE
                   AND location_id IN (60, 80)
                   AND historico != 'yes'
                   AND state = 'done')
            ) <= 0 THEN 0
            ELSE (
                (SELECT COALESCE(SUM(product_qty), 0)
                 FROM stock_move 
                 WHERE product_id = p.id 
                   AND date <= CURRENT_DATE
                   AND location_dest_id IN (60, 80)
                   AND historico != 'yes'
                   AND state = 'done')
                -
                (SELECT COALESCE(SUM(product_qty), 0)
                 FROM stock_move 
                 WHERE product_id = p.id 
                   AND date <= CURRENT_DATE
                   AND location_id IN (60, 80)
                   AND historico != 'yes'
                   AND state = 'done')
            )
        END AS "Cantidad",
 
        COALESCE(pcps.standard_price_real, 0.0) AS "Precio de coste",
 
        CASE 
            WHEN (
                (SELECT COALESCE(SUM(product_qty), 0)
                 FROM stock_move 
                 WHERE product_id = p.id 
                   AND date <= CURRENT_DATE
                   AND location_dest_id IN (60, 80)
                   AND historico != 'yes'
                   AND state = 'done')
                -
                (SELECT COALESCE(SUM(product_qty), 0)
                 FROM stock_move 
                 WHERE product_id = p.id 
                   AND date <= CURRENT_DATE
                   AND location_id IN (60, 80)
                   AND historico != 'yes'
                   AND state = 'done')
            ) <= 0 THEN 0
            ELSE (
                (
                    (SELECT COALESCE(SUM(product_qty), 0)
                     FROM stock_move 
                     WHERE product_id = p.id 
                       AND date <= CURRENT_DATE
                       AND location_dest_id IN (60, 80)
                       AND historico != 'yes'
                       AND state = 'done')
                    -
                    (SELECT COALESCE(SUM(product_qty), 0)
                     FROM stock_move 
                     WHERE product_id = p.id 
                       AND date <= CURRENT_DATE
                       AND location_id IN (60, 80)
                       AND historico != 'yes'
                       AND state = 'done')
                ) * COALESCE(pcps.standard_price_real, 0.0)
            )
        END AS "Precio de coste total"
 
    FROM 
        product_product p
    INNER JOIN 
        product_cost_price_sede pcps ON (pcps.product_id = p.id AND pcps.sede_id = 6)
 
    WHERE 
        p.id IN (
            SELECT id FROM product_product 
            WHERE type = 'product' AND active AND NOT pack_ok
        )
 
    ORDER BY 
        p.id;
    """
    
    # 4. Conectar a la base de datos, ejecutar la consulta y obtener los datos
    try:
        with psycopg2.connect(**db_params) as conn:
            with conn.cursor() as cur:
                cur.execute(query)
                resultados = cur.fetchall()
                headers = [desc[0] for desc in cur.description]  # Encabezados de columnas
    except Exception as e:
        print(f"Error al conectar o ejecutar la consulta: {e}")
        sys.exit(1)
        
    if not resultados:
        print("No se obtuvieron resultados de la consulta.")
        return
    else:
        print(f"Se obtuvieron {len(resultados)} filas de la consulta.")
    
    # 5. Abrir el archivo Excel. Si no existe, se crea un nuevo libro e incluye los encabezados.
    try:
        book = load_workbook(file_path)
        sheet = book.active
    except FileNotFoundError:
        print(f"No se encontró el archivo '{file_path}'. Se creará uno nuevo.")
        book = Workbook()
        sheet = book.active
        sheet.append(headers)
    
    # 6. Evitar duplicados usando la columna "Referencia" (sexta columna, índice 5)
    existing_refs = {row[5] for row in sheet.iter_rows(min_row=2, values_only=True)}
    for row in resultados:
        if row[5] not in existing_refs:
            sheet.append(row)
            new_row_index = sheet.max_row
            # Copiar formato de la fila anterior para mantener consistencia (si existe una fila previa)
            if new_row_index > 1:
                for col in range(1, sheet.max_column + 1):
                    source_cell = sheet.cell(row=new_row_index - 1, column=col)
                    target_cell = sheet.cell(row=new_row_index, column=col)
                    target_cell.font = copy.copy(source_cell.font)
                    target_cell.fill = copy.copy(source_cell.fill)
                    target_cell.border = copy.copy(source_cell.border)
                    target_cell.alignment = copy.copy(source_cell.alignment)
    
    # 7. Actualizar la referencia de la tabla si existe en el libro
    # Se asume que la tabla se llama "Productos"
    if "Productos" in sheet.tables:
        tabla = sheet.tables["Productos"]
        max_row = sheet.max_row
        max_col = sheet.max_column
        last_col_letter = get_column_letter(max_col)
        new_ref = f"A1:{last_col_letter}{max_row}"
        tabla.ref = new_ref
        print(f"Tabla 'Productos' actualizada a rango: {new_ref}")
    else:
        print("No se encontró la tabla 'Productos'. Se conservará el formato actual, pero no se actualizará la referencia de la tabla.")
    
    # 8. Guardar el archivo Excel con el nombre basado en la fecha y hora
    book.save(file_path)
    print(f"Archivo guardado en '{file_path}'.")
    
    # 9. Si se ejecuta en GitHub Actions, se establece la salida para usar el nombre del archivo en el flujo (OneDrive)
    # GitHub Actions utiliza la variable de entorno GITHUB_OUTPUT para establecer outputs
    github_output = os.environ.get("GITHUB_OUTPUT")
    if github_output:
        with open(github_output, "a") as gh_out:
            print(f"file_path={file_path}", file=gh_out)
    
if __name__ == '__main__':
    main()
